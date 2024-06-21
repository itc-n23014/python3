import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

n_price = {'Garlic': 3.07,'Celery': 1.19,'Lemon': 1.27 }

#価格の更新
for r in range(2, sheet.max_row): #最初の行をストップ
    name = sheet.cell(row=r, column=1).value 
    if name in n_price:  
        sheet.cell(row=r, column=2).value = n_price[name]

wb.save('updatedProduceSales.xlsx')  
