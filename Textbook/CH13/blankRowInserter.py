import openpyxl,sys

if len(sys.argv) < 4:
    sys.exit("""
使い方: python3 script.py N M [スプレッドシート]
  N = 何行ごとに区切るか
  M = 空行の数
""")

n,m = int(sys.argv[1]),int(sys.argv[2])
src = sys.argv[3]

wb = openpyxl.load_workbook(src)
sheet = wb.active

new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

for row in range(1, sheet.max_row + 1):
    new_row = row + m if row >= n else row
    for col in range(1, sheet.max_column + 1):
        new_sheet.cell(column=col, row=new_row).value = sheet.cell(column=col, row=row).value

new_wb.save(src + '.ins.xlsx')
