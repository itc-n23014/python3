import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active

s_wb = openpyxl.Workbook()
s_sheet = s_wb.active

data = list(sheet.iter_rows(values_only=True))
[s_sheet.cell(row=i+1, column=j+1, value=value) for i, row in enumerate(zip(*data)) for j, value in enumerate(row)]

s_wb.save('example.swap.xlsx')

