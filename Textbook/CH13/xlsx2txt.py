from openpyxl import load_workbook

input_filename = "input.xlsx"

ws = load_workbook(input_filename).active

[open(file_name, 'w').close() for file_name in (row[0] for row in ws.iter_rows(min_row=1, max_col=1, values_only=True)) if file_name]

print("列上の名前のファイルを作成しました.")

